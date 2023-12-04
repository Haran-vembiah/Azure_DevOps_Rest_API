"""
Module for parsing and creating xlsx files based on templates or in their raw form.

Cell styles are not handled very well by this module. Best option for formatting is to use the template file with
conditional formatting, because even existing styles are not preserved when writing to an xlsx file.

To parse an xlsx file, use the parse_xlsx() function. It returns a dictionary where the keys are the names of the
worksheets in the xlsx file and the values are lists of lists, where each list represents a row in the worksheet.

To write an xlsx file, use the write_to_xlsx() function. It takes a dictionary as input, where the keys are the names of
the worksheets in the output file and the values are lists of lists, where each list represents a row in the worksheet.
The function also takes a template file as input, which is used to create the output file. See the docstring of the
write_to_xlsx() function for more details.

This module uses the openpyxl package mostly because the alternatives, xlrd, xlwt, xlwings etc.,
are either not maintained anymore, or require Excel to be installed on the system.

One alternative to openpyxl is poi, a Python wrapper for the Java library Apache POI.
https://github.com/ryanwang520/poi
In fact, we should probably switch to poi as soon as any additional functionality is needed, like proper cell styling.

(Check this answer on StackOverflow for more info on the differences between xlwings and openpyxl:
https://stackoverflow.com/a/58331928/12446690)
The openpyxl package is documented here: https://openpyxl.readthedocs.io/en/stable/
Of particular interest are the Workbook, Worksheet, Cell and MergedCell classes.

Workbook
    https://openpyxl.readthedocs.io/en/stable/api/openpyxl.workbook.workbook.html#openpyxl.workbook.workbook.Workbook

Worksheet
    https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.worksheet.html#openpyxl.worksheet.worksheet.Worksheet

Cell
    https://openpyxl.readthedocs.io/en/stable/api/openpyxl.cell.cell.html#openpyxl.cell.cell.Cell

MergedCell
    Used to detect merged cells in a worksheet and get the value of the only non-merged cell in the range.
    https://openpyxl.readthedocs.io/en/stable/api/openpyxl.cell.cell.html#openpyxl.cell.cell.MergedCell

MergedCellRange
    Used to get the coordinates of all the cells in a merged cell range. These coordinates are used to
    check if a cell is a part of a merged cell range.
    https://openpyxl.readthedocs.io/en/stable/api/openpyxl.worksheet.merge_range.html#openpyxl.worksheet.merge_range.MergedCellRange

Conditional Formatting
    As mentioned above, the best option for formatting is to use the template file with conditional formatting.
    When a template file is used, and a worksheet is copied into the output file, the conditional formatting is not
    copied automatically, it needs to be copied manually.
    https://openpyxl.readthedocs.io/en/stable/formatting.html?#conditional-formatting

"""

import logging
from pathlib import Path
from typing import Union

import openpyxl as px
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import TYPE_FORMULA, MergedCell


def determine_merged_range_values(ws: Worksheet) -> dict:
    """
    This function checks all the merged cells in a worksheet and returns a dictionary where the keys are the coordinates
    of the cells in the merged cell ranges and the value of the only non-merged cell in the range is copied to all the
    cells in the range.

    :param ws: A worksheet object from the openpyxl package
    :return: A dictionary as described above
    """
    merged_ranges = ws.merged_cells.ranges
    merged_range_values = {}
    for merged_range in merged_ranges:
        merged_range_value = None
        for cell_coords in merged_range.cells:
            cell = ws.cell(cell_coords[0], cell_coords[1])
            if not isinstance(cell, MergedCell):
                if cell.data_type == TYPE_FORMULA:
                    merged_range_value = f"FORMULA[{cell.value.strip('=')}]"
                else:
                    merged_range_value = cell.value
                continue
        for cell_coords in merged_range.cells:
            merged_range_values[cell_coords] = merged_range_value
    return merged_range_values


def _parse_ws(ws: Worksheet) -> list:
    """
    This function takes a worksheet object from the openpyxl package and returns a list of lists where each list
    represents a row in the worksheet. Merged cells are handled by copying the value of the only non-merged cell in the
    merged cell range to all the cells.
    :param ws: A worksheet object from the openpyxl package
    :return: List of lists as described above
    """
    ws_data = []
    logging.info(f"Worksheet: {ws.title}")
    merged_range_values = determine_merged_range_values(ws)
    for row in ws.rows:
        row_data = []
        for cell in row:
            normalized_cell_value = None
            if isinstance(cell, MergedCell):
                normalized_cell_value = merged_range_values[(cell.row, cell.column)]
            else:
                normalized_cell_value = cell.value
            if cell.data_type == TYPE_FORMULA:
                row_data.append(f"FORMULA[{normalized_cell_value.strip('=')}]")
            else:
                row_data.append(normalized_cell_value)
        ws_data.append(row_data)
    return ws_data


def parse_xlsx(path: Union[str, Path]) -> dict:
    """
    The parse_xlsx function takes a path to an XLSX file and returns a dictionary of the data in the file.
    Merged cells are handled by copying the value of the only non-merged cell in the merged cell range to all the cells.
    :param path: The path to the XLSX file
    :return: A dictionary where the keys are the worksheet names and the values
    are the data in the worksheets as a list of lists.
    """
    xlsx_data = {}
    if not isinstance(path, Path):
        path = Path(path)
    wb = px.load_workbook(path)
    for ws in wb.worksheets:
        xlsx_data[ws.title] = _parse_ws(ws)
    return xlsx_data


def write_to_xlsx(path: Union[str, Path], data: dict, template_file: Union[str, Path] = None):
    """
    The write_to_xlsx function takes a dictionary of data and writes it to an XLSX file.
    The input dictionary's keys are the worksheet names and the values are the data to be written to the worksheet.
    Optionally, a template XLSX file can be passed to the function. If a template file is passed, the function will use
    the worksheets in the template file as the worksheets in the output file.
    The input dictionary may contain metadata. In this case, the actual data content must be in a key called "data".
    The metadata must be in a key called "meta". The metadata must be a dictionary and may contain the following keys:
    - "template_sheet": The name of the worksheet in the template file to use as the output worksheet.
        The name of the sheet in the output file will NOT be the same as the name of the sheet in the template file.
        Instead, the name of the sheet in the output file will be the same as the key in the input dictionary.
        !!! If a template_sheet name is defined, but is not found in the template file, the function will raise an error.
    - "start_cell": Must be a tuple of integers, where the first item is the number of rows,
        the second item is the number of columns to skip before writing the data to the worksheet.

    Notes on the template file: The openpyxl package does not support copying worksheets from one workbook to another,
    so the template file is used as a reference for the formatting of the output file. The template file is not modified.
    Openpyxl ignores the formatting of the cells in the template file, if the cells don't contain actual values,
    so to define a properly formatted worksheet, the best practice is to use Conditional Formatting in the templates.

    :param path: The path to the output XLSX file. If the parent directory does not exist, it will be created.
    :param data: A dictionary of data to be written to the output XLSX file. The structure is described above.
    :param template_file: An optional path to a template XLSX file. The details are described above.
    :return: No return value. If no errors are raised, the output file is written to the specified path. (Hopefully.)
    """
    ws_list = []
    if not isinstance(path, Path):
        logging.info(f"A string was passed as the output XLSX path, converting to Path object: {path}")
        path = Path(path)
    if not path.parent.exists():
        logging.info(f"Creating parent directory for output XLSX file: {path}")
        path.parent.mkdir(parents=True)

    if template_file:
        if not isinstance(template_file, Path):
            logging.info(f"A string was passed as the Template XLSX path, converting to Path object: {template_file}")
            template_file = Path(template_file)
        if not template_file.exists():
            raise FileNotFoundError(f"Template XLSX file not found: {template_file}")
        else:
            wb = px.load_workbook(template_file)
            ws_list = set([ws.title for ws in wb.worksheets])
            template_list = set([meta['meta']['template_sheet'] for meta in data.values() if 'meta' in meta])
            logging.info(f"Available worksheets in template file: {ws_list}")
            logging.info(f"Worksheets to be created from templates: {template_list}")
            if not template_list.issubset(ws_list):
                raise ValueError(f"Template worksheet names not found in template file: {template_list - ws_list}")
    else:
        wb = px.Workbook()
        wb.remove(wb.active)

    logging.info(f"Writing data to XLSX file: {path}")
    for ws_name in data.keys():
        row_offset, col_offset = 1, 1
        sheet_meta = None
        if isinstance(data[ws_name], dict):
            logging.info(f"Found metadata for worksheet: {ws_name}")
            if 'meta' in data[ws_name]:
                sheet_meta = data[ws_name].pop('meta')
            if 'data' in data[ws_name]:
                ws_data = data[ws_name].pop('data')
            else:
                raise ValueError(f"No data found to be inserted into worksheet: {ws_name}")
        else:
            ws_data = data[ws_name]

        if sheet_meta:
            if 'template_sheet' in sheet_meta:
                if sheet_meta['template_sheet'] not in ws_list:
                    raise ValueError(f"The template sheet {sheet_meta['template_sheet']} "
                                     f"was not found in the template workbook {template_file}")
                else:
                    ws = wb.copy_worksheet(wb.get_sheet_by_name(sheet_meta['template_sheet']))
                    ws.title = ws_name
                    ws.conditional_formatting = \
                        wb.get_sheet_by_name(sheet_meta['template_sheet']).conditional_formatting
            else:
                raise ValueError(f"The template sheet {sheet_meta['template_sheet']} "
                                 f"was not found in the template workbook {template_file}")
            if 'start_cell' in sheet_meta:
                row_offset, col_offset = sheet_meta['start_cell']
        else:
            ws = wb.create_sheet(ws_name)
        logging.info(f"Writing data to worksheet: {ws_name}")
        for row_num, row in enumerate(ws_data):
            for col_num, cell_data in enumerate(row):
                active_cell = ws.cell(row_num + row_offset, col_num + col_offset, cell_data)
                if active_cell.has_style:
                    logging.info(f"The cell style is: {active_cell._style}")
    # Cleanup: deleting all the worksheets which are not in the list of worksheets to be created
    for ws in wb.worksheets:
        if ws.title not in data.keys():
            logging.info(f"Deleting worksheet: {ws.title}")
            wb.remove(ws)
    logging.info(f"Saving XLSX file: {path}")
    wb.save(path)
